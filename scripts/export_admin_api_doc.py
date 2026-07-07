"""导出管理员端接口文档为 Excel。"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADERS = [
    "编号",
    "页面",
    "功能描述",
    "请求参数",
    "响应数据",
    "类型",
    "请求url路径",
    "备注",
    "说明",
    "完成情况",
]

ROWS = [
    [
        "A-01",
        "管理员登录页",
        "管理员登录",
        "Body JSON：\nphone：手机号，String，必填\npassword：密码，String，必填",
        '成功：{"access_token":"xxx",\n'
        '"token_type":"bearer",\n'
        '"user":\n'
        '{"id":1,\n'
        '"username":"admin",\n'
        '"role":"admin"}}\n'
        '失败：401 账号或密码错误\n'
        '403 {"detail":"账号已被禁用"}',
        "POST",
        "/api/v1/auth/login",
        "前端需校验 user.role === admin",
        "与用户端共用登录接口，通过角色区分",
        "已实现",
    ],
    [
        "A-02",
        "管理仪表盘",
        "获取运维总览数据",
        "无",
        '成功：{"health":{...},\n'
        '"cards":{...},\n'
        '"vector_snapshots":{...},\n'
        '"trends":{...},\n'
        '"error_clusters":[...],\n'
        '"top_users":[...]}',
        "GET",
        "/api/v1/admin/overview",
        "health 含系统健康评分",
        "管理首页核心数据",
        "已实现",
    ],
    [
        "A-03",
        "运维总览",
        "查询系统暂停状态",
        "无",
        '成功：{"paused": false}\n'
        '失败：403 {"detail":"无管理员权限"}',
        "GET",
        "/api/v1/admin/system/pause",
        "paused=true 表示全局暂停",
        "",
        "已实现",
    ],
    [
        "A-04",
        "运维总览",
        "设置系统暂停/恢复",
        "Body JSON：\npaused：是否暂停，Boolean，必填",
        '成功：{"paused": true,\n'
        '"message": "全局暂停系统"}\n'
        '或 {"paused": false,\n'
        '"message": "恢复系统运行"}',
        "POST",
        "/api/v1/admin/system/pause",
        "高风险操作，写入审计日志",
        "",
        "已实现",
    ],
    [
        "A-05",
        "运维总览",
        "获取向量库快照趋势",
        "Query：\ndays：天数，int，默认 7\nlimit：条数上限，int，默认 168",
        '成功：{"items":[...],\n'
        '"latest":{...},\n'
        '"series":[100,120,...],\n'
        '"days":7}',
        "GET",
        "/api/v1/admin/vector/snapshots",
        "items 含 total_vectors、storage_mb、health_score 等",
        "",
        "已实现",
    ],
    [
        "A-06",
        "运维总览",
        "系统健康检查",
        "无",
        '成功：{"status":"ok","redis":true} 或 degraded',
        "GET",
        "/api/v1/system/health",
        "无需管理员权限（公开）",
        "兼容接口",
        "已实现",
    ],
    [
        "A-07",
        "运维总览",
        "运营统计数据",
        "无",
        '成功：{"total_uploaded":0,\n'
        '"total_parsed":0,\n'
        '"total_qa_calls":0,\n'
        '"total_reports":0,\n'
        '"total_graphs":0,\n'
        '"total_users":0,\n'
        '"active_users":0,\n'
        '"vector_db_total":0}',
        "GET",
        "/api/v1/system/operation-stats",
        "",
        "",
        "已实现",
    ],
    [
        "A-08",
        "运维总览",
        "近7日每日统计",
        "无",
        '成功：{"dates":["..."],"upload":[...],"parse":[...],"qa":[...],"report":[...]}',
        "GET",
        "/api/v1/system/daily-stats",
        "数组与 dates 一一对应",
        "",
        "已实现",
    ],
    [
        "A-09",
        "用户管理页",
        "查询用户列表",
        "Query：\nkeyword：关键词，String，可选\nstatus：状态，String，可选\nsort_by：排序字段，String，可选\nsort_order：排序方向，String，默认 desc",
        "成功：用户数组，含 id/username/name/email/phone/role/status/quota 等\n失败：403 无管理员权限",
        "GET",
        "/api/v1/admin/users",
        "邮箱/手机号自动脱敏；最多返回 500 条",
        "",
        "已实现",
    ],
    [
        "A-10",
        "用户管理页",
        "获取用户详情",
        "Path：\nuser_id：用户 ID，int，必填",
        '成功：{"user":{...},\n'
        '"audit_logs":[...]}\n'
        '失败：404 {"detail":"用户不存在"}',
        "GET",
        "/api/v1/admin/users/{user_id}/detail",
        "audit_logs 含最近 10 条操作记录",
        "",
        "已实现",
    ],
    [
        "A-11",
        "用户管理页",
        "新增用户",
        "Body JSON：\nusername：用户名，String，3-50 字符，必填\npassword：密码，String，6-128 字符，必填\nphone：手机号，String，必填\nname/email/role/status/quota：可选",
        '成功：{"id": 2,\n'
        '"message": "已新增用户"}\n'
        '失败：400 {"detail":"账号、邮箱或手机号已存在"}',
        "POST",
        "/api/v1/admin/users",
        "role：researcher/admin；status：active/disabled",
        "",
        "已实现",
    ],
    [
        "A-12",
        "用户管理页",
        "更新用户信息",
        "Path：user_id：用户 ID，int，必填\nBody JSON（可选）：name/email/phone/status/role/quota",
        '成功：{"message": "已更新"}\n失败：404 用户不存在',
        "PUT",
        "/api/v1/admin/users/{user_id}",
        "前端用于修改配额、启用/禁用用户",
        "",
        "已实现",
    ],
    [
        "A-13",
        "用户管理页",
        "删除用户",
        "Path：user_id：用户 ID，int，必填",
        '成功：{"message": "已删除"}\n失败：400 不能删除当前登录管理员\n404 用户不存在',
        "DELETE",
        "/api/v1/admin/users/{user_id}",
        "",
        "",
        "已实现",
    ],
    [
        "A-14",
        "用户管理页",
        "重置用户密码",
        "Path：user_id：用户 ID，int，必填\nQuery：new_password：新密码，String，必填",
        '成功：{"message": "已重置"}\n失败：404 用户不存在',
        "POST",
        "/api/v1/admin/users/{user_id}/reset-password",
        "",
        "",
        "已实现",
    ],
    [
        "A-15",
        "用户管理页",
        "导出用户 CSV",
        "无",
        "成功：CSV 文件流（Content-Type: text/csv）",
        "GET",
        "/api/v1/admin/users/export",
        "邮箱/手机号脱敏导出",
        "",
        "已实现",
    ],
    [
        "A-16",
        "模型管理页",
        "获取模型配置列表",
        "无",
        "成功：模型数组，含 id/model_type/model_name/scenario/is_primary/is_active 等",
        "GET",
        "/api/v1/model-configs",
        "api_key 脱敏返回",
        "",
        "已实现",
    ],
    [
        "A-17",
        "模型管理页",
        "获取模型监控数据",
        "无",
        '成功：{"date":"...","summary":{...},"items":[...],"model_count":4,...}',
        "GET",
        "/api/v1/admin/models/monitor",
        "含近 7 日调用趋势",
        "",
        "已实现",
    ],
    [
        "A-18",
        "模型管理页",
        "获取 LLM 场景列表",
        "无",
        '成功：[{"id":"parse","name":"文献结构化提取"},{"id":"qa","name":"实时知识问答"},...]',
        "GET",
        "/api/v1/admin/models/scenarios",
        "",
        "",
        "已实现",
    ],
    [
        "A-19",
        "模型管理页",
        "新增模型配置",
        "Body JSON：\nmodel_type：parse/vector/reranker/llm，必填\nmodel_name：必填\nversion/api_endpoint/config/is_active/scenario/is_primary：可选",
        '成功：{"id": 5}\n失败：400 校验失败或未配置密钥',
        "POST",
        "/api/v1/model-configs",
        "LLM 类型需配置 api_key",
        "",
        "已实现",
    ],
    [
        "A-20",
        "模型管理页",
        "更新模型配置",
        "Path：model_id：模型 ID，int，必填\nBody JSON（可选）：model_type/model_name/version/api_endpoint/config/is_active/scenario/is_primary",
        '成功：{"message": "updated"}\n失败：404 模型不存在',
        "PUT",
        "/api/v1/model-configs/{model_id}",
        "config 中 api_key 留空则保留原值",
        "",
        "已实现",
    ],
    [
        "A-21",
        "模型管理页",
        "获取模型详情",
        "Path：model_id：模型 ID，int，必填",
        "成功：含 7 日统计、config（脱敏）等\n失败：404 模型不存在",
        "GET",
        "/api/v1/model-configs/{model_id}/detail",
        "",
        "",
        "已实现",
    ],
    [
        "A-22",
        "模型管理页",
        "查询当前生效 LLM 运行时",
        "Query：scenario：场景，String，默认 qa",
        '成功（已配置）：{"configured":true,"scenario":"qa",...}\n成功（未配置）：{"configured":false,"message":"..."}',
        "GET",
        "/api/v1/model-configs/active-llm",
        "",
        "",
        "已实现",
    ],
    [
        "A-23",
        "任务管理页",
        "分页查询任务列表",
        "Query：\npage：页码，int，默认 1\npage_size：每页条数，int，默认 20，最大 100\nstatus：状态筛选，String，可选",
        '成功：{"items":[...],\n'
        '"total":100,\n'
        '"page":1,\n'
        '"page_size":20,\n'
        '"stats":{...}}',
        "GET",
        "/api/v1/admin/tasks",
        "status：queued/running/completed/failed/cancelled/pending",
        "",
        "已实现",
    ],
    [
        "A-24",
        "任务管理页",
        "批量重试任务",
        "Body JSON：\ntask_ids：任务 ID 数组，Array[int]，1-100 个，必填",
        '成功：{"retried_task_ids": [1, 3, 5]}',
        "POST",
        "/api/v1/admin/tasks/batch-retry",
        "仅 failed/cancelled 状态可重试",
        "",
        "已实现",
    ],
    [
        "A-25",
        "任务管理页",
        "批量终止任务",
        "Body JSON：\ntask_ids：任务 ID 数组，Array[int]，1-100 个，必填",
        '成功：{"cancelled_task_ids": [2, 4]}',
        "POST",
        "/api/v1/admin/tasks/batch-cancel",
        "已完成/失败/已取消的不会被终止",
        "",
        "已实现",
    ],
    [
        "A-26",
        "任务管理页",
        "终止单个任务",
        "Path：task_id：任务 ID，int，必填",
        '成功：{"message": "已终止"}\n失败：400 当前状态不允许终止\n404 任务不存在',
        "POST",
        "/api/v1/tasks/{task_id}/cancel",
        "",
        "",
        "已实现",
    ],
    [
        "A-27",
        "任务管理页",
        "重试单个任务",
        "Path：task_id：任务 ID，int，必填",
        '成功：{"message": "已重新加入队列", "task_id": 1}\n失败：400 只有失败或已终止任务可重试',
        "POST",
        "/api/v1/tasks/{task_id}/retry",
        "",
        "",
        "已实现",
    ],
    [
        "A-28",
        "任务管理页",
        "调整任务优先级",
        "Path：task_id：任务 ID，int，必填\nBody JSON：priority：优先级，int，0-9，必填",
        '成功：{"message": "优先级已调整", "task_id": 1, "priority": 3}',
        "PUT",
        "/api/v1/tasks/{task_id}/priority",
        "数值越小越优先",
        "",
        "已实现",
    ],
    [
        "A-29",
        "任务管理页",
        "获取任务失败统计",
        "无",
        '成功：{"failed_total": 10, "by_reason": [{"reason": "...", "count": 5}]}',
        "GET",
        "/api/v1/tasks/failure-stats",
        "按 error_log 聚合",
        "",
        "已实现",
    ],
    [
        "A-30",
        "任务管理页",
        "获取调度器配置",
        "无",
        '成功：{"max_concurrent_tasks":4,"per_user_concurrent":2,"timeout_seconds":300,"alert_rules":{},"backup_engine_config":{}}',
        "GET",
        "/api/v1/tasks/scheduler-config",
        "",
        "",
        "已实现",
    ],
    [
        "A-31",
        "任务管理页",
        "保存调度器配置",
        "Body JSON：\nmax_concurrent_tasks/per_user_concurrent/timeout_seconds/alert_rules/backup_engine_config",
        '成功：{"message": "已保存"}',
        "PUT",
        "/api/v1/tasks/scheduler-config",
        "",
        "",
        "已实现",
    ],
    [
        "A-32",
        "任务管理页",
        "任务状态 SSE 推送",
        "Path：task_id：任务 ID，int，必填",
        "成功：text/event-stream 流式推送任务状态变更",
        "GET",
        "/api/v1/tasks/{task_id}/events",
        "实时监听任务进度",
        "",
        "已实现",
    ],
    [
        "A-33",
        "审计中心",
        "分页查询审计日志",
        "Query：\npage/size/user_id/risk_flag/keyword/start_at/end_at",
        '成功：{"items":[...],\n'
        '"total":100,\n'
        '"page":1,\n'
        '"size":20}',
        "GET",
        "/api/v1/admin/logs/audit",
        "仅展示终端用户行为；risk_flag：0正常/1中风险/2高风险",
        "",
        "已实现",
    ],
    [
        "A-34",
        "审计中心",
        "查询审计日志（旧版）",
        "Query：module：模块名，String，可选",
        "成功：审计日志数组（最多 500 条）",
        "GET",
        "/api/v1/audit-logs",
        "已废弃，建议使用 A-33",
        "deprecated",
        "已实现",
    ],
    [
        "A-35",
        "运维总览",
        "获取向量库实时统计",
        "无",
        '成功：{"total_vectors":1000,"storage_mb":50.2,"health_score":100,...}',
        "GET",
        "/api/v1/vector-store/stats",
        "同时写入快照表",
        "",
        "已实现",
    ],
    [
        "A-36",
        "运维总览",
        "获取备份列表",
        "无",
        "成功：备份数组，含 id/backup_type/backup_time/file_size_mb/status 等",
        "GET",
        "/api/v1/vector-store/backups",
        "最多 100 条",
        "",
        "已实现",
    ],
    [
        "A-37",
        "运维总览",
        "创建向量库备份",
        "无",
        '成功：{"id":2,"status":"completed","file_location":"...",...}\n失败：500 备份失败',
        "POST",
        "/api/v1/vector-store/backups",
        "高风险操作，写入审计",
        "",
        "已实现",
    ],
    [
        "A-38",
        "运维总览",
        "恢复向量库备份",
        "Body JSON：\nbackup_id：备份 ID，int，必填",
        '成功：{"restore_task_id":1,"status":"completed",...}\n失败：404 备份不存在\n500 恢复失败',
        "POST",
        "/api/v1/vector-store/restore",
        "高风险操作",
        "",
        "已实现",
    ],
]

COLUMN_WIDTHS = [8, 14, 16, 36, 42, 8, 42, 24, 20, 10]


def _style_header(ws) -> None:
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[1].height = 28


def _style_rows(ws, row_count: int) -> None:
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill("solid", fgColor="F2F2F2")

    for row_idx in range(2, row_count + 1):
        ws.row_dimensions[row_idx].height = 72
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = alt_fill


def export_excel(output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "管理员端接口文档"

    _style_header(ws)

    for row in ROWS:
        ws.append(row)

    _style_rows(ws, len(ROWS) + 1)

    for idx, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"

    # 附录 sheet
    appendix = wb.create_sheet("通用约定")
    appendix.append(["项目", "说明"])
    appendix_rows = [
        ["Base URL", "/api/v1"],
        ["认证方式", "Authorization: Bearer {access_token}"],
        ["权限要求", "除 A-01、A-06 外，均需管理员角色（role=admin）"],
        ["role 枚举", "researcher 科研人员 / admin 管理员"],
        ["status 枚举", "active 正常 / disabled 禁用"],
        ["任务 status", "pending / queued / running / completed / failed / cancelled"],
        ["model_type", "parse 解析 / vector 向量 / reranker 重排 / llm 大语言模型"],
        ["错误响应格式", '{"detail": "错误描述"} 或 {"detail": [{...}]}（422 校验错误）'],
    ]
    for row in appendix_rows:
        appendix.append(row)

    appendix.column_dimensions["A"].width = 18
    appendix.column_dimensions["B"].width = 70
    for row in appendix.iter_rows(min_row=1, max_row=len(appendix_rows) + 1, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


if __name__ == "__main__":
    out = Path(__file__).resolve().parents[1] / "docs" / "管理员端接口文档.xlsx"
    export_excel(out)
    print(f"已导出: {out}")
